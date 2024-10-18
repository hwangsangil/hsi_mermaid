import openpyxl
import re
from datetime import datetime
import pytz
from datetime import datetime

# 시스템 이름 패턴과 클래스 매핑
# 기본값으로 "central" 클래스를 사용하므로 별도로 명시하지 않음
class_mappings = {
    "security": [
        "Privacy", "Log system", "DB Access Control"
    ],
    "local-rectangle": [
        "MINI Online", "Helpdesk Ticketing", "Data Warehouse", "Motorrad Online",
        "BMW Vantage", "CSM KR", "E-tax tool", "BMW Online", "Intranet",
        "Cloud Shared Service", "RDC Operation System", "Local Cloud Data Hub KR"
    ],
    "local-circle": [
        "My DMS", "Vehicle Tracing System"
    ],
}

# mermaid_code 시작 - 각 도형 class 별 모양 정의
mermaid_code = '---\n \
config:\n \
    layout: elk\n \
    elk:\n \
        mergeEdges: true\n \
        nodePlacementStrategy: NETWORK_SIMPLEX\n \
---\n \
graph LR\n \
classDef local-circle fill:#0166B1,stroke:#333,stroke-width:2px,color:#ffffff;\n \
classDef local-rectangle fill:#0166B1,stroke:#333,stroke-width:2px,color:#ffffff;\n \
classDef security fill:#000000,stroke:#333,stroke-width:2px,color:#ffffff;\n \
classDef central fill:#6F6F6F,stroke:#333,stroke-width:2px,color:#ffffff;\n \
classDef highlight stroke:red,stroke-width:6px,color:#ffffff;\n'

# Load the workbook and select the active sheet
workbook = openpyxl.load_workbook('./excel/IT_Product___Sub_Product_SWP_3048__Integration_Mar.xlsx')
# workbook = openpyxl.load_workbook('IT_Product___Sub_Product_SWP_3048__Integration_Mar.xlsx')
sheet = workbook.active

# Initialize a dictionary to store the Information Flow data
information_flows = {}

# Set to store all unique systems
systems = set()

# Dictionary to map full system names to simplified names
system_mapping = {}

# Iterate through the rows in the sheet
for row in sheet.iter_rows(min_row=4):  # Starting from row 4 as IDs start from B4 and Information Flow from C4
    id_cell = row[1]  # Column B is the second column, index starts from 0
    flow_cell = row[2]  # Column C is the third column
    if id_cell.value and flow_cell.value:  # Check if both cells are not empty
        # Split the Information Flow into source and destination
        source, destination = flow_cell.value.split('>>')
        source = source.strip()
        destination = destination.strip()
        
        # Add source and destination to the systems set
        systems.add(source)
        systems.add(destination)
        
        # Store in dictionary
        information_flows[id_cell.value.strip()] = (source, destination)

# Generate all possible two-letter combinations
def generate_alphabet_combinations():
    alphabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    combinations = []
    for letter1 in alphabet:
        for letter2 in alphabet:
            combinations.append(f"{letter1}{letter2}")
    return combinations

# Use the function to generate combinations
alphabet_combinations = generate_alphabet_combinations()

# Sort the systems and alphabet_combinations lists (시스템 이름 순서 고정)
sorted_systems = sorted(systems)

# Assign a unique two-letter combination to each system
for i, system in enumerate(sorted_systems):
    simplified_name = alphabet_combinations[i]  # Use the generated combinations
    system_mapping[system] = simplified_name



# 시스템별 클래스 정보를 저장할 딕셔너리 초기화
system_classes = {}

# 시스템 이름을 검사하여 클래스 정보 업데이트
for system in systems:
    # 클래스 매핑을 통해 시스템 클래스 결정
    assigned = False  # 시스템에 클래스가 할당되었는지 추적
    for class_name, patterns in class_mappings.items():
        if any(pattern in system for pattern in patterns):
            system_classes[system] = class_name
            assigned = True
            break  # 매칭되는 첫 번째 클래스에 할당하고 반복 중단
    if not assigned:
        # 어떤 패턴과도 매치되지 않으면 "central" 클래스 할당
        system_classes[system] = "central"
# print(system_classes)
# Define each system uniquely with simplified names
for system, simplified_name in system_mapping.items():
    if system_classes[system] == 'local-circle':
        mermaid_code += f'    {simplified_name}(("{system}"))\n'
    elif system_classes[system] =='local-rectangle':
        mermaid_code += f'    {simplified_name}(["{system}"])\n'
    else:
        mermaid_code += f'    {simplified_name}["{system}"]\n'
# Generate flows between systems using simplified names
for id, (source, destination) in information_flows.items():
    mermaid_code += f'    {system_mapping[source]} -->|{id}| {system_mapping[destination]}\n'

for system, class_name in system_classes.items():
    simplified_name = system_mapping[system]  # 시스템의 간소화된 이름 가져오기
    mermaid_code += f'    class {simplified_name} {class_name};\n'

# Get current date and time in KST (Korea Standard Time)

# Set the timezone to KST
kst = pytz.timezone('Asia/Seoul')
current_datetime = datetime.now(tz=kst).strftime("%Y-%m-%d %H:%M:%S")
print("지금시간 : ", current_datetime)

# Print or save the mermaid code
print(mermaid_code)
# HTML 파일 경로
html_file_path = './html/mermaid.html'
# html_file_path = 'mermaid.html'

# 파일 읽기
with open(html_file_path, 'r', encoding='utf-8') as file:
    html_content = file.read()

# <div class="mermaid"> 태그 내용 업데이트
new_html_content = re.sub(
    r'(<div class="mermaid">).*?(</div>)', 
    r'\1' + mermaid_code + r'\2', 
    html_content, 
    flags=re.DOTALL
)

# Append current date and time right before the </body> tag
new_html_content = re.sub(
    r'(</body>)', 
    f'<p style="text-align: right;">last updated: {current_datetime}</p>\n\\1', 
    new_html_content
)

# 변경된 내용으로 파일 다시 쓰기
with open(html_file_path, 'w', encoding='utf-8') as file:
    file.write(new_html_content)